import openpyxl

workbook = openpyxl.load_workbook("ApocolypseFoodPrep.xlsx")
worksheet = workbook['Sheet1']

#dictionary for counting the various store name
productPerStore = {}
# totalSalesPerStore = {}
# for row in range(2, worksheet.max_row+1):
#     sales = row[2].value

#min_row specifies the minimum row to begin reading from in order to avoid the heading
#max_row specifies the maximum row in the worksheet
for row in worksheet.iter_rows(min_row=2, max_row=worksheet.max_row):

    if row[0].value in productPerStore:
        productPerStore[row[0].value] += 1
    else:
        productPerStore[row[0].value] = 1
print(productPerStore)